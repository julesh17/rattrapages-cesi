"""
Application Streamlit - Gestion des rattrapages étudiants (v4)
Lancer avec : streamlit run rattrapages_app_v4.py
Dépendances : pip install streamlit pandas openpyxl
"""

import io
import re
import base64
import streamlit as st
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG PAGE ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Rattrapages – Tableau de bord",
    page_icon="🎓",
    layout="wide",
)

# ─── STYLES CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700;800&family=DM+Mono:wght@400;500&display=swap');

* { font-family: 'DM Sans', sans-serif; }

[data-testid="stAppViewContainer"] {
    background: #f4f6fb;
}
[data-testid="stSidebar"] {
    background: #1e1b4b !important;
}
[data-testid="stSidebar"] * {
    color: #e0e7ff !important;
}
[data-testid="stSidebar"] .stMarkdown h3 {
    color: #a5b4fc !important;
    font-size: 0.72rem !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: 1.5px !important;
    margin-top: 1.4rem !important;
    margin-bottom: 0.4rem !important;
    border-bottom: 1px solid rgba(165,180,252,0.2) !important;
    padding-bottom: 4px !important;
}
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] .stCheckbox label span,
[data-testid="stSidebar"] .stSelectbox label {
    color: #c7d2fe !important;
    font-size: 0.82rem !important;
}
[data-testid="stSidebar"] [data-testid="stMultiSelect"] span {
    background: #3730a3 !important;
    color: #e0e7ff !important;
}
[data-testid="stSidebar"] [data-testid="baseButton-secondary"] {
    background: #3730a3 !important;
    color: white !important;
    border: 1px solid #4f46e5 !important;
}

.hero {
    background: linear-gradient(135deg, #1e1b4b 0%, #312e81 50%, #4338ca 100%);
    border-radius: 16px; padding: 1.8rem 2.2rem; margin-bottom: 1.5rem;
    color: white; box-shadow: 0 8px 32px rgba(67,56,202,0.3);
    display: flex; align-items: center; gap: 1.5rem;
}
.hero-icon { font-size: 2.8rem; }
.hero h1 { margin: 0; font-size: 1.7rem; font-weight: 800; letter-spacing: -0.5px; }
.hero p  { margin: 0.2rem 0 0; opacity: 0.7; font-size: 0.88rem; }

.step-card {
    background: white;
    border-radius: 14px;
    padding: 1.4rem 1.6rem;
    margin-bottom: 1rem;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    border: 1px solid #e8eaf6;
}
.step-header {
    display: flex; align-items: center; gap: 10px;
    margin-bottom: 0.8rem;
}
.step-num {
    background: #4f46e5; color: white; border-radius: 8px;
    width: 28px; height: 28px; display: inline-flex;
    align-items: center; justify-content: center;
    font-size: 0.82rem; font-weight: 800; flex-shrink: 0;
}
.step-title {
    font-size: 0.95rem; font-weight: 700; color: #1e1b4b; margin: 0;
}

.badge {
    display: inline-block; font-weight: 700; font-size: 0.78rem;
    padding: 2px 9px; border-radius: 20px; margin: 1px;
}
.badge-A   { background:#d1fae5; color:#065f46; border:1px solid #6ee7b7; }
.badge-B   { background:#dbeafe; color:#1e3a8a; border:1px solid #93c5fd; }
.badge-C   { background:#fef3c7; color:#78350f; border:1px solid #fcd34d; }
.badge-D   { background:#fee2e2; color:#7f1d1d; border:1px solid #fca5a5; }
.badge-ABS { background:#ffedd5; color:#7c2d12; border:1px solid #fb923c; }
.badge-VAL  { background:#d1fae5; color:#065f46; border:1px solid #6ee7b7; font-size:0.72rem; }
.badge-NVAL { background:#fee2e2; color:#7f1d1d; border:1px solid #fca5a5; font-size:0.72rem; }
.badge-COMP { background:#e0e7ff; color:#3730a3; border:1px solid #a5b4fc; font-size:0.72rem; }

.legend-row { display:flex; gap:10px; align-items:center; flex-wrap:wrap; margin-bottom:0.5rem; }
.legend-item { display:flex; align-items:center; gap:6px; font-size:0.82rem; font-weight:600; color:#374151; }
.dot { width:12px; height:12px; border-radius:50%; display:inline-block; }
.dot-A{background:#10b981;} .dot-B{background:#3b82f6;}
.dot-C{background:#f59e0b;} .dot-D{background:#ef4444;} .dot-ABS{background:#f97316;}

.section-title {
    font-size:1rem; font-weight:700; color:#1e1b4b;
    margin-bottom:0.5rem; display:flex; align-items:center; gap:8px;
}

.stat-pill {
    display: inline-flex; align-items: center; gap: 5px;
    background: #f0f4ff; border: 1px solid #c7d2fe;
    border-radius: 20px; padding: 3px 12px;
    font-size: 0.78rem; font-weight: 700; color: #3730a3;
}
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTES ─────────────────────────────────────────────────────────────────
GRADE_VALUES = {"A": 5, "B": 4, "C": 2, "D": 1}
RAT_VALS     = {"C", "D", "ABS"}

# ─── FONCTIONS UTILITAIRES ───────────────────────────────────────────────────────

def badge_html(val):
    if pd.isna(val) or str(val).strip() == "":
        return ""
    v = str(val).strip()
    css = f"badge badge-{v}" if v in ("A","B","C","D","ABS") else "badge"
    return f'<span class="{css}">{v}</span>'


def split_name(personne: str):
    if "," in personne:
        p = personne.split(",", 1)
        return p[1].strip().title(), p[0].strip().title()
    tokens = personne.strip().split()
    if len(tokens) >= 2:
        return " ".join(tokens[1:]).title(), tokens[0].title()
    return personne.title(), ""


def short_name(col: str) -> str:
    return re.sub(r"^Eval\s*-\s*", "", col).strip()


def copy_button_html(text: str, uid: str) -> str:
    b64  = base64.b64encode(text.encode("utf-8")).decode()
    safe = re.sub(r"[^a-zA-Z0-9_]", "_", uid)
    return (
        f'<button id="btn_{safe}" onclick="'
        f'var t=atob(\'{b64}\');'
        f'navigator.clipboard.writeText(t).then(function(){{'
        f'var b=document.getElementById(\'btn_{safe}\');'
        f'b.innerText=\'✅ Copié !\';b.style.background=\'#d1fae5\';b.style.color=\'#065f46\';'
        f'setTimeout(function(){{b.innerText=\'📋 Copier\';'
        f'b.style.background=\'#ede9fe\';b.style.color=\'#4f46e5\';}},2000);}});" '
        f'style="width:100%;padding:8px 12px;border:1px solid #c4b5fd;border-radius:8px;'
        f'background:#ede9fe;color:#4f46e5;font-weight:600;font-size:0.85rem;cursor:pointer;">'
        f'📋 Copier</button>'
    )


def generate_email(prenom, nom, matieres, tutoyer):
    liste = "\n".join(f"  • {m}" for m in matieres)
    if tutoyer:
        return (f"Bonjour {prenom},\n\n"
                "Nous t'informons que tu es concerné(e) par des rattrapages dans les matières suivantes :\n\n"
                f"{liste}\n\n"
                "Nous t'invitons donc à te présenter aux sessions de rattrapage "
                "qui te seront communiquées prochainement.\n\n"
                "N'hésite pas à nous contacter si tu as des questions.\n\n"
                "Bien cordialement,\nL'équipe pédagogique")
    return (f"Bonjour {prenom} {nom},\n\n"
            "Nous vous informons que vous êtes concerné(e) par des rattrapages dans les matières suivantes :\n\n"
            f"{liste}\n\n"
            "Nous vous invitons donc à vous présenter aux sessions de rattrapage "
            "qui vous seront communiquées prochainement.\n\n"
            "N'hésitez pas à nous contacter si vous avez des questions.\n\n"
            "Bien cordialement,\nL'équipe pédagogique")


# ─── LOGIQUE COMPENSATION ────────────────────────────────────────────────────────

@st.cache_data
def load_ue_structure(rn_bytes: bytes, semestre: int) -> dict:
    df = pd.read_excel(io.BytesIO(rn_bytes))
    df = df[df["Semestre Unite Enseignement"] == semestre]
    df = df.dropna(subset=["Libelle Element Evaluable", "Coefficient Element Evaluable"])
    df = df[df["Coefficient Element Evaluable"] > 0]
    struct = {}
    for _, row in df.iterrows():
        ue = row["Libelle Unite Enseignement"]
        struct.setdefault(ue, []).append({
            "element": row["Libelle Element Evaluable"],
            "coeff":   float(row["Coefficient Element Evaluable"]),
        })
    return struct


def match_element(elem_name: str, grades_row: dict):
    el = elem_name.lower().strip()
    for key, val in grades_row.items():
        kl = key.lower().strip()
        if el in kl or kl in el:
            return str(val).strip() if val else None
    stop = {'', 'de', 'la', 'le', 'les', 'du', 'et', 'en', 'un', 'une', 'a', 'au'}
    ew = set(re.split(r'[\s\-:,\(\)]+', el)) - stop
    for key, val in grades_row.items():
        kw = set(re.split(r'[\s\-:,\(\)]+', key.lower())) - stop
        if len(ew & kw) >= 3:
            return str(val).strip() if val else None
    return None


def compute_ue_result(grades_row: dict, ue_elements: list) -> dict:
    elements_data = []
    total_coeff = weighted_sum = 0.0
    missing = False

    for ei in ue_elements:
        mention = match_element(ei["element"], grades_row)
        coeff   = ei["coeff"]
        if mention is None or mention == "ABS":
            elements_data.append({"element": ei["element"], "coeff": coeff,
                                   "mention": mention or "—", "value": None})
            missing = True
        elif mention in GRADE_VALUES:
            v = GRADE_VALUES[mention]
            weighted_sum += v * coeff
            total_coeff  += coeff
            elements_data.append({"element": ei["element"], "coeff": coeff,
                                   "mention": mention, "value": v})
        else:
            elements_data.append({"element": ei["element"], "coeff": coeff,
                                   "mention": mention, "value": None})
            missing = True

    if total_coeff == 0:
        return {"mention": None, "weighted_avg": None, "elements": elements_data,
                "validated": False, "compensation": False, "missing": True}

    avg = weighted_sum / total_coeff
    if   avg > 4.6: mention_ue, validated = "A", True
    elif avg > 3.6: mention_ue, validated = "B", True
    elif avg > 1.6: mention_ue, validated = "C", False
    else:           mention_ue, validated = "D", False

    has_cd = any(e["mention"] in ("C","D") for e in elements_data if e["value"] is not None)
    compensation = validated and has_cd

    return {"mention": mention_ue, "weighted_avg": round(avg, 3),
            "elements": elements_data, "validated": validated,
            "compensation": compensation, "missing": missing}


def is_compensated_for(skey: str, col_short: str, ue_results: dict) -> bool:
    for result in ue_results.get(skey, {}).values():
        if result["validated"] and result["compensation"]:
            for e in result["elements"]:
                es, el = col_short.lower(), e["element"].lower()
                if es in el or el in es:
                    return True
    return False


# ─── EXPORT EXCEL ────────────────────────────────────────────────────────────────

def make_excel(display_df, eval_disp_cols, student_ue_results, use_comp, all_students_df) -> bytes:
    thin   = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    HDR      = PatternFill("solid", fgColor="1E1B4B")
    HDR2     = PatternFill("solid", fgColor="312E81")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    c_center = Alignment(horizontal="center", vertical="center")
    c_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    FILLS = {
        "A":   PatternFill("solid", fgColor="D1FAE5"),
        "B":   PatternFill("solid", fgColor="DBEAFE"),
        "C":   PatternFill("solid", fgColor="FEF3C7"),
        "D":   PatternFill("solid", fgColor="FEE2E2"),
        "ABS": PatternFill("solid", fgColor="FFEDD5"),
    }
    F_RED   = PatternFill("solid", fgColor="FEE2E2")
    F_GREEN = PatternFill("solid", fgColor="D1FAE5")
    F_COMP  = PatternFill("solid", fgColor="E0E7FF")
    F_EMPTY = PatternFill("solid", fgColor="F9FAFB")
    F_HDR_ROW = PatternFill("solid", fgColor="EEF2FF")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Onglet 1 : Vue globale (TOUTES personnes × TOUTES matières) ──────────
        # On travaille sur l'ensemble des étudiants (pas seulement filtrés)
        vg_df = all_students_df.copy()
        vg_df = vg_df[["Prénom", "Nom"] + [c for c in eval_disp_cols if c in vg_df.columns]]
        vg_df = vg_df.sort_values(["Nom", "Prénom"]).reset_index(drop=True)
        vg_df.to_excel(writer, index=False, sheet_name="Vue globale")
        ws_vg = writer.sheets["Vue globale"]

        # En-têtes
        for cell in ws_vg[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = HDR
            cell.alignment = Alignment(horizontal="center", vertical="bottom",
                                       text_rotation=45, wrap_text=False)
            cell.border = border
        ws_vg.row_dimensions[1].height = 90
        ws_vg.column_dimensions["A"].width = 14
        ws_vg.column_dimensions["B"].width = 14
        for i in range(3, ws_vg.max_column + 1):
            ws_vg.column_dimensions[get_column_letter(i)].width = 7

        for row in ws_vg.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = c_center
                if cell.value in FILLS:
                    cell.fill = FILLS[cell.value]
                elif cell.value is None or str(cell.value).strip() == "":
                    cell.fill = F_EMPTY
            # Nom / Prénom alignés à gauche
            row[0].alignment = c_left
            row[1].alignment = c_left

        # ── Onglet 2 : Rattrapages synthèse ──────────────────────────────────────
        rows_rat = []
        for _, row in all_students_df.iterrows():
            skey = f"{row['Prénom']} {row['Nom']}"
            rat, comp = [], []
            for col in eval_disp_cols:
                if col not in row:
                    continue
                g = str(row.get(col, "")).strip()
                if g in RAT_VALS:
                    if use_comp and is_compensated_for(skey, col, student_ue_results):
                        comp.append(col)
                    else:
                        rat.append(col)
            rows_rat.append({
                "Prénom": row["Prénom"], "Nom": row["Nom"],
                "Matières en rattrapage":           ", ".join(rat)  or "—",
                "Nb rattrapages":                   len(rat),
                "Matières compensées (dispensées)": ", ".join(comp) or "—" if use_comp else "N/A",
                "Nb compensées":                    len(comp) if use_comp else 0,
            })
        df_rat = (pd.DataFrame(rows_rat)
                  .sort_values(["Nb rattrapages", "Nom"], ascending=[False, True]))
        df_rat.to_excel(writer, index=False, sheet_name="Rattrapages")
        ws2 = writer.sheets["Rattrapages"]
        for cell in ws2[1]:
            cell.font, cell.fill, cell.alignment = hdr_font, HDR, c_center
            cell.border = border
        ws2.row_dimensions[1].height = 38
        for row_cells in ws2.iter_rows(min_row=2):
            nb_rat  = row_cells[3].value or 0
            nb_comp = row_cells[5].value or 0
            if nb_rat > 0:
                fill = F_RED
            elif nb_comp > 0:
                fill = F_COMP
            else:
                fill = F_GREEN
            for cell in row_cells:
                cell.border = border
                cell.fill   = fill
                cell.alignment = c_left
            row_cells[3].alignment = c_center
            row_cells[5].alignment = c_center
        for col_letter, w in zip("ABCDEF", [14, 14, 45, 12, 45, 12]):
            ws2.column_dimensions[col_letter].width = w

        # ── Onglet 3 : Mentions (étudiants filtrés) ───────────────────────────────
        display_df.to_excel(writer, index=False, sheet_name="Mentions (filtrés)")
        ws_m = writer.sheets["Mentions (filtrés)"]
        for cell in ws_m[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = HDR2
            cell.alignment = Alignment(horizontal="center", vertical="bottom",
                                       text_rotation=45, wrap_text=False)
            cell.border = border
        ws_m.row_dimensions[1].height = 90
        ws_m.column_dimensions["A"].width = 14
        ws_m.column_dimensions["B"].width = 14
        for i in range(3, ws_m.max_column + 1):
            ws_m.column_dimensions[get_column_letter(i)].width = 7
        for row in ws_m.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = c_center
                if cell.value in FILLS:
                    cell.fill = FILLS[cell.value]
        for row in ws_m.iter_rows(min_row=2):
            row[0].alignment = c_left
            row[1].alignment = c_left

        # ── Onglet 4 : Résultats UE (si compensation active) ─────────────────────
        if use_comp and student_ue_results:
            rows_ue = []
            for _, row in all_students_df.iterrows():
                skey = f"{row['Prénom']} {row['Nom']}"
                for ue_name, res in student_ue_results.get(skey, {}).items():
                    if res["weighted_avg"] is None:
                        continue
                    if res["validated"] and res["compensation"]:
                        statut = "Validée par compensation"
                    elif res["validated"]:
                        statut = "Validée"
                    else:
                        statut = "Non validée"
                    rows_ue.append({
                        "Prénom": row["Prénom"], "Nom": row["Nom"],
                        "UE": ue_name,
                        "Mention UE": res["mention"] or "—",
                        "Moy. pondérée": res["weighted_avg"],
                        "Statut": statut,
                    })
            if rows_ue:
                pd.DataFrame(rows_ue).to_excel(writer, index=False, sheet_name="Résultats UE")
                ws3 = writer.sheets["Résultats UE"]
                for cell in ws3[1]:
                    cell.font, cell.fill, cell.alignment = hdr_font, HDR, c_center
                    cell.border = border
                ws3.row_dimensions[1].height = 35
                S_FILLS = {"Validée par compensation": F_COMP,
                           "Validée": F_GREEN, "Non validée": F_RED}
                for row_cells in ws3.iter_rows(min_row=2):
                    statut_val = row_cells[5].value if len(row_cells) > 5 else None
                    fill = S_FILLS.get(statut_val, PatternFill())
                    for cell in row_cells:
                        cell.border, cell.fill = border, fill
                        cell.alignment = c_left
                    for idx in [3, 4]:
                        if len(row_cells) > idx:
                            row_cells[idx].alignment = c_center
                for col_letter, w in zip("ABCDEF", [14, 14, 40, 10, 12, 22]):
                    ws3.column_dimensions[col_letter].width = w

    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
# ─── INTERFACE ──────────────────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="hero">
  <div class="hero-icon">🎓</div>
  <div>
    <h1>Gestion des rattrapages</h1>
    <p>Importez les notes, configurez les options, consultez les récaps et générez les convocations.</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# ─── SIDEBAR : CONFIGURATION ────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### 📂 Fichiers")
    uploaded_notes = st.file_uploader("Fichier de notes (.xlsx)", type=["xlsx"], key="notes",
                                       label_visibility="collapsed")
    st.caption("Fichier de notes (.xlsx)")
    uploaded_rn = st.file_uploader("Référentiel cahier des charges (.xlsx)", type=["xlsx"], key="rn",
                                    label_visibility="collapsed",
                                    help="Fichier RN — structure UE, éléments évaluables, coefficients")
    st.caption("Référentiel RN (.xlsx) — optionnel")

    st.markdown("### 📅 Semestre")
    semestre = st.selectbox("Semestre", options=[5, 6, 7, 8], index=2, label_visibility="collapsed")

    st.markdown("### ⚙️ Matières")
    # Ces widgets seront peuplés après chargement du fichier
    sidebar_excl_placeholder = st.empty()
    sidebar_toggle_placeholder = st.empty()
    sidebar_abs_placeholder = st.empty()

    st.markdown("### 🔍 Filtres")
    st.markdown("""
    <div class="legend-row" style="margin-bottom:0.3rem;">
      <span class="legend-item"><span class="dot dot-A"></span>A Admis</span>
      <span class="legend-item"><span class="dot dot-B"></span>B Bien</span>
      <span class="legend-item"><span class="dot dot-C"></span>C Ajourné</span>
      <span class="legend-item"><span class="dot dot-D"></span>D Ajourné+</span>
      <span class="legend-item"><span class="dot dot-ABS"></span>ABS</span>
    </div>""", unsafe_allow_html=True)
    selected_grades = st.multiselect(
        "Mentions à inclure", ["A","B","C","D"], default=["C","D"],
        label_visibility="collapsed")
    group_filter = st.radio("Vue", [
        "Tous",
        "Admis uniquement (A/B)",
        "À rattraper uniquement (C/D)",
    ], index=2, label_visibility="collapsed")

    st.markdown("### ⚖️ Compensation UE")
    use_comp = st.toggle(
        "Activer les compensations UE",
        value=(uploaded_rn is not None),
        help="Nécessite le fichier référentiel.")

    st.markdown("### ✉️ Mails")
    tutoyer = st.toggle("Tutoyer les étudiants", value=False)
    if st.button("🔄 Rafraîchir les mails", use_container_width=True,
                 help="Remet à zéro les textes des mails après changement de filtre ou d'option"):
        for key in list(st.session_state.keys()):
            if key.startswith("mail_") or key == "recap_classe":
                del st.session_state[key]
        st.rerun()

# ─── VÉRIFICATION FICHIER ────────────────────────────────────────────────────────

if uploaded_notes is None:
    st.markdown("""
    <div class="step-card" style="text-align:center;padding:2.5rem;">
      <div style="font-size:3rem;margin-bottom:1rem;">📂</div>
      <div style="font-size:1.1rem;font-weight:700;color:#1e1b4b;margin-bottom:0.5rem;">
        Commencez par importer votre fichier de notes
      </div>
      <div style="font-size:0.88rem;color:#6b7280;">
        Utilisez le panneau latéral gauche pour charger le fichier Excel (.xlsx)
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

notes_bytes = uploaded_notes.read()
try:
    raw_df = pd.read_excel(io.BytesIO(notes_bytes))
except Exception as e:
    st.error(f"Impossible de lire le fichier de notes : {e}")
    st.stop()

if "Personne" not in raw_df.columns:
    st.error("Colonne 'Personne' introuvable.")
    st.stop()

all_eval_cols = [c for c in raw_df.columns if str(c).startswith("Eval")]
if not all_eval_cols:
    st.error("Aucune colonne commençant par 'Eval' trouvée.")
    st.stop()

# ─── MATIÈRES À INCLURE (dans la sidebar) ────────────────────────────────────────

GLOBAL_KW  = "Préparation à la certification (Global exam)"
has_global = any(GLOBAL_KW.lower() in c.lower() for c in all_eval_cols)

short_map = {c: short_name(c) for c in all_eval_cols}
short_inv = {v: k for k, v in short_map.items()}

with sidebar_toggle_placeholder:
    exclude_global = st.toggle("Exclure « Global exam »", value=False) if has_global else False

default_excl = (
    [short_map[c] for c in all_eval_cols if GLOBAL_KW.lower() in c.lower()]
    if exclude_global else []
)
with sidebar_excl_placeholder:
    excluded_short = st.multiselect(
        "Exclure des matières", options=list(short_map.values()),
        default=default_excl, label_visibility="collapsed",
        placeholder="Sélectionner matières à exclure…")

with sidebar_abs_placeholder:
    absent_toggle = st.toggle("Cellule vide = ABS (convoqué)", value=False)

excluded_full = {short_inv[s] for s in excluded_short}
if exclude_global:
    excluded_full |= {c for c in all_eval_cols if GLOBAL_KW.lower() in c.lower()}

eval_cols = [c for c in all_eval_cols if c not in excluded_full]
if not eval_cols:
    st.error("Toutes les matières sont exclues.")
    st.stop()

# ─── NETTOYAGE DATA ──────────────────────────────────────────────────────────────

working = raw_df[["Personne"] + eval_cols].copy()
working = working[working[eval_cols].notna().any(axis=1)].reset_index(drop=True)
working.insert(0, "Prénom", working["Personne"].apply(lambda x: split_name(x)[0]))
working.insert(1, "Nom",    working["Personne"].apply(lambda x: split_name(x)[1]))
working.drop(columns=["Personne"], inplace=True)

col_rename     = {c: short_name(c) for c in eval_cols}
display_df     = working.rename(columns=col_rename)
eval_disp_cols = list(col_rename.values())

if absent_toggle:
    active_cols = [c for c in eval_disp_cols if display_df[c].notna().any()]
    for col in active_cols:
        display_df[col] = display_df[col].apply(
            lambda v: "ABS" if (pd.isna(v) or str(v).strip() == "") else v)

# ─── COMPENSATIONS UE ────────────────────────────────────────────────────────────

ue_structure       = {}
student_ue_results = {}

if use_comp:
    if uploaded_rn is None:
        st.warning("⚠️ Importez le fichier référentiel pour activer les compensations.")
        use_comp = False
    else:
        try:
            rn_bytes     = uploaded_rn.read()
            ue_structure = load_ue_structure(rn_bytes, semestre)

            for _, row in display_df.iterrows():
                skey   = f"{row['Prénom']} {row['Nom']}"
                grades = {
                    col: (str(row[col]).strip() if pd.notna(row[col]) else None)
                    for col in eval_disp_cols
                }
                student_ue_results[skey] = {
                    ue: compute_ue_result(grades, elems)
                    for ue, elems in ue_structure.items()
                }
        except Exception as e:
            st.error(f"Erreur référentiel : {e}")
            use_comp = False

# ─── FILTRE ─────────────────────────────────────────────────────────────────────

def apply_filter(df):
    cols = eval_disp_cols
    if selected_grades:
        with_abs = selected_grades + (["ABS"] if any(g in selected_grades for g in ["C","D"]) else [])
        df = df[df[cols].isin(with_abs).any(axis=1)]
    if group_filter == "Admis uniquement (A/B)":
        df = df[df[cols].isin(["A","B"]).any(axis=1) & ~df[cols].isin(list(RAT_VALS)).any(axis=1)]
    elif group_filter == "À rattraper uniquement (C/D)":
        df = df[df[cols].isin(list(RAT_VALS)).any(axis=1)]
    return df


filtered_df = apply_filter(display_df.copy())

# ─── BANDEAU RÉSUMÉ ──────────────────────────────────────────────────────────────

nb_excl = len(excluded_full)
comp_status = f"⚖️ {len(ue_structure)} UE" if (use_comp and ue_structure) else "⚖️ Compensation off"
st.markdown(
    f'<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:1.2rem;">'
    f'<span class="stat-pill">👥 {len(working)} étudiants</span>'
    f'<span class="stat-pill">📚 {len(eval_cols)} matières{" ("+str(nb_excl)+" exclue(s))" if nb_excl else ""}</span>'
    f'<span class="stat-pill">📅 Semestre {semestre}</span>'
    f'<span class="stat-pill">{comp_status}</span>'
    f'<span class="stat-pill">🔍 {len(filtered_df)} résultat(s)</span>'
    f'</div>',
    unsafe_allow_html=True
)

# ═══════════════════════════════════════════════════════════════════════════════
# ─── ONGLETS PRINCIPAUX ─────────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📋 Tableau des résultats",
    "📊 Récap par matière",
    "✉️ Mails de convocation",
    "🗓️ Créneaux parallèles",
    "📥 Export Excel",
])

# ═══════════════════════════════════════════════════════════════════════════════
# ─── TAB 1 : TABLEAU DES RÉSULTATS ──────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

with tab1:
    if filtered_df.empty:
        st.warning("Aucun étudiant ne correspond aux filtres.")
    else:
        headers = "".join(
            f'<th style="background:#1e1b4b;color:white;padding:8px 10px;'
            f'font-size:0.78rem;text-align:center;white-space:nowrap;">{h}</th>'
            for h in filtered_df.columns)
        rows_html = ""
        for i, (_, row) in enumerate(filtered_df.iterrows()):
            cells = ""
            for col in filtered_df.columns:
                v = row[col]
                if col in ("Prénom","Nom"):
                    cells += (f'<td style="padding:6px 10px;font-weight:600;'
                              f'white-space:nowrap;font-size:0.85rem;">{v}</td>')
                else:
                    cells += f'<td style="text-align:center;padding:4px 8px;">{badge_html(v)}</td>'
            bg = "#f8fafc" if i % 2 == 0 else "white"
            rows_html += f'<tr style="background:{bg}">{cells}</tr>'
        st.markdown(
            f'<div style="overflow-x:auto;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,0.08);">'
            f'<table style="width:100%;border-collapse:collapse;font-family:sans-serif;">'
            f'<thead><tr>{headers}</tr></thead><tbody>{rows_html}</tbody></table></div>',
            unsafe_allow_html=True)

    # Résultats UE (si compensation active)
    if use_comp and ue_structure and not filtered_df.empty:
        st.markdown("---")
        st.markdown('<div class="section-title">⚖️ Résultats par UE avec compensations</div>',
                    unsafe_allow_html=True)
        st.markdown(
            '<p style="font-size:0.82rem;color:#6b7280;margin-bottom:1rem;">'
            'A si moy&gt;4,6 · B si moy&gt;3,6 · C (non validée) si moy&gt;1,6 · D sinon'
            ' — <em>A=5 · B=4 · C=2 · D=1</em></p>',
            unsafe_allow_html=True)

        for _, row in filtered_df.iterrows():
            skey       = f"{row['Prénom']} {row['Nom']}"
            ue_results = student_ue_results.get(skey, {})
            if not ue_results:
                continue
            non_val = sum(1 for r in ue_results.values() if not r["validated"] and not r["missing"])
            comped  = sum(1 for r in ue_results.values() if r["validated"] and r["compensation"])
            icon    = "🔴" if non_val else "🟢"

            with st.expander(f"{icon} {skey} — {non_val} UE non validée(s) · {comped} compensation(s)"):
                for ue_name, res in ue_results.items():
                    if res["weighted_avg"] is None:
                        continue
                    if res["missing"]:
                        bg2, bd2 = "#f8fafc", "#e5e7eb"
                    elif res["validated"] and res["compensation"]:
                        bg2, bd2 = "#e0e7ff", "#a5b4fc"
                    elif res["validated"]:
                        bg2, bd2 = "#d1fae5", "#6ee7b7"
                    else:
                        bg2, bd2 = "#fee2e2", "#fca5a5"

                    mbadge = (f'<span class="badge badge-{res["mention"]}">{res["mention"]}</span>'
                              if res["mention"] else "")
                    if res["missing"]:
                        sbadge = '<span class="badge" style="background:#f3f4f6;color:#6b7280;border:1px solid #d1d5db;">Données manquantes</span>'
                    elif res["validated"] and res["compensation"]:
                        sbadge = '<span class="badge badge-COMP">✓ Validée par compensation</span>'
                    elif res["validated"]:
                        sbadge = '<span class="badge badge-VAL">✓ Validée</span>'
                    else:
                        sbadge = '<span class="badge badge-NVAL">✗ Non validée</span>'

                    elems_html = ""
                    for e in res["elements"]:
                        em = e["mention"]
                        eb = (f'<span class="badge badge-{em}">{em}</span>'
                              if em in ("A","B","C","D","ABS")
                              else f'<span style="color:#9ca3af;font-size:0.8rem;">{em}</span>')
                        elems_html += (
                            f'<div style="display:flex;justify-content:space-between;align-items:center;'
                            f'padding:3px 0;border-bottom:1px solid #f3f4f6;">'
                            f'<span style="font-size:0.78rem;color:#374151;flex:1;">{e["element"]}</span>'
                            f'<span style="font-size:0.75rem;color:#6b7280;margin:0 8px;">coeff {int(e["coeff"])}</span>'
                            f'{eb}</div>')
                    st.markdown(
                        f'<div style="background:{bg2};border:1.5px solid {bd2};border-radius:10px;'
                        f'padding:12px 16px;margin-bottom:8px;">'
                        f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;flex-wrap:wrap;">'
                        f'<span style="font-weight:700;font-size:0.88rem;flex:1;">{ue_name}</span>'
                        f'{mbadge}'
                        f'<span style="font-size:0.82rem;color:#6b7280;">moy. {res["weighted_avg"]:.2f}</span>'
                        f'{sbadge}</div>'
                        f'<div style="border-top:1px solid {bd2};padding-top:8px;">{elems_html}</div></div>',
                        unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# ─── TAB 2 : RÉCAP PAR MATIÈRE ──────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

with tab2:
    recap_rows = []
    if not filtered_df.empty:
        for col in eval_disp_cols:
            ec, ed, ea, ec_comp, ed_comp = [], [], [], [], []
            for _, row in filtered_df.iterrows():
                skey  = f"{row['Prénom']} {row['Nom']}"
                grade = str(row.get(col, "")).strip()
                if grade == "C":
                    if use_comp and is_compensated_for(skey, col, student_ue_results):
                        ec_comp.append(skey)
                    else:
                        ec.append(skey)
                elif grade == "D":
                    if use_comp and is_compensated_for(skey, col, student_ue_results):
                        ed_comp.append(skey)
                    else:
                        ed.append(skey)
                elif grade == "ABS":
                    ea.append(skey)

            total_rat = len(ec) + len(ed) + len(ea)
            total_all = total_rat + len(ec_comp) + len(ed_comp)
            if total_all > 0:
                recap_rows.append({
                    "Matière": col,
                    "eleves_c": ec,       "nb_c": len(ec),
                    "eleves_d": ed,       "nb_d": len(ed),
                    "eleves_abs": ea,     "nb_abs": len(ea),
                    "eleves_c_comp": ec_comp, "eleves_d_comp": ed_comp,
                    "total_rat": total_rat, "total_all": total_all,
                })

    if not recap_rows:
        st.info("Aucune matière avec C ou D pour les étudiants sélectionnés.")
    else:
        recap_rows.sort(key=lambda x: -x["total_rat"])
        max_rat = max(r["total_rat"] for r in recap_rows) or 1
        nb_comp_total = sum(len(r["eleves_c_comp"]) + len(r["eleves_d_comp"]) for r in recap_rows)
        st.markdown(
            f'<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:1rem;">'
            f'<span class="stat-pill">📚 {len(recap_rows)} matière(s)</span>'
            f'<span class="stat-pill" style="background:#fef3c7;border-color:#fcd34d;color:#78350f;">C : {sum(r["nb_c"] for r in recap_rows)}</span>'
            f'<span class="stat-pill" style="background:#fee2e2;border-color:#fca5a5;color:#7f1d1d;">D : {sum(r["nb_d"] for r in recap_rows)}</span>'
            f'<span class="stat-pill" style="background:#ffedd5;border-color:#fb923c;color:#7c2d12;">ABS : {sum(r["nb_abs"] for r in recap_rows)}</span>'
            + (f'<span class="stat-pill" style="background:#e0e7ff;border-color:#a5b4fc;color:#3730a3;">⚖️ {nb_comp_total} compensé(s)</span>' if nb_comp_total else "")
            + f'</div>',
            unsafe_allow_html=True)

        def pill(name, mention, compensated=False):
            if compensated:
                bg2, fg2, bd2, label = "#e0e7ff","#3730a3","#a5b4fc",f"{name} ⚖️"
            else:
                bg2, fg2, bd2 = {"C":("#fef3c7","#78350f","#fcd34d"),
                                  "D":("#fee2e2","#7f1d1d","#fca5a5"),
                                  "ABS":("#ffedd5","#7c2d12","#fb923c")}[mention]
                label = name
            return (f'<span style="display:inline-block;background:{bg2};color:{fg2};'
                    f'border:1px solid {bd2};border-radius:20px;padding:1px 10px;'
                    f'font-size:0.76rem;font-weight:600;margin:2px;">{label}</span>')

        for i, r in enumerate(recap_rows):
            bg = "#f8fafc" if i % 2 == 0 else "white"
            bw = int(r["total_rat"] / max_rat * 100)
            bc = int(r["nb_c"] / max(r["total_rat"],1) * bw)
            bd_w = int(r["nb_d"] / max(r["total_rat"],1) * bw)
            ba = bw - bc - bd_w
            bar = (f'<div style="display:flex;height:6px;border-radius:4px;overflow:hidden;'
                   f'width:{bw}%;min-width:4px;margin-top:5px;">'
                   f'<div style="flex:{bc} 0 0;background:#f59e0b;"></div>'
                   f'<div style="flex:{bd_w} 0 0;background:#ef4444;"></div>'
                   f'<div style="flex:{ba} 0 0;background:#f97316;"></div></div>')
            b_c   = f'<span class="badge badge-C">{r["nb_c"]}</span>'     if r["nb_c"]   else ""
            b_d   = f'<span class="badge badge-D">{r["nb_d"]}</span>'     if r["nb_d"]   else ""
            b_abs = f'<span class="badge badge-ABS">{r["nb_abs"]}</span>' if r["nb_abs"] else ""
            nb_c2 = len(r["eleves_c_comp"]) + len(r["eleves_d_comp"])
            b_comp = f'<span class="badge badge-COMP">⚖️ {nb_c2} compensé(s)</span>' if nb_c2 else ""
            all_pills = (
                "".join(pill(e,"C")      for e in r["eleves_c"])
              + "".join(pill(e,"D")      for e in r["eleves_d"])
              + "".join(pill(e,"ABS")    for e in r["eleves_abs"])
              + "".join(pill(e,"C",True) for e in r["eleves_c_comp"])
              + "".join(pill(e,"D",True) for e in r["eleves_d_comp"])
            )
            st.markdown(
                f'<div style="background:{bg};border-radius:10px;padding:10px 16px;'
                f'margin-bottom:6px;box-shadow:0 1px 4px rgba(0,0,0,0.05);border:1px solid #e8eaf6;">'
                f'<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;">'
                f'<div style="flex:1;min-width:200px;">'
                f'<span style="font-weight:700;font-size:0.88rem;">{r["Matière"]}</span>{bar}</div>'
                f'<div style="display:flex;gap:5px;align-items:center;flex-wrap:wrap;">'
                f'{b_c}{b_d}{b_abs}{b_comp}'
                f'<span style="font-size:0.78rem;color:#6b7280;">/ {r["total_rat"]}</span></div></div>'
                f'<div style="margin-top:7px;line-height:2;">{all_pills}</div></div>',
                unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# ─── TAB 3 : MAILS DE CONVOCATION ───────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

with tab3:
    tkey = "tu" if tutoyer else "vous"

    # ── Mail récap classe ──────────────────────────────────────────────────────
    st.markdown('<div class="section-title">📣 Mail récap à la classe</div>', unsafe_allow_html=True)

    if not recap_rows:
        st.info("Aucune donnée à afficher — consultez d'abord l'onglet Récap.")
    else:
        lines = ["Bonjour à tous,", "",
                 "Voici le récapitulatif des convocations aux rattrapages par matière :", ""]
        for r in recap_rows:
            if r["total_rat"] == 0:
                continue
            students = r["eleves_c"] + r["eleves_d"] + r["eleves_abs"]
            lines.append(f"• {r['Matière']} : {', '.join(students)}")
        lines += ["", "Les étudiants concernés sont invités à se présenter aux sessions de rattrapage "
                  "dont les modalités leur seront communiquées prochainement.",
                  "", "Bien cordialement,", "L'équipe pédagogique"]

        recap_txt = "\n".join(lines)
        edited_recap = st.text_area("Mail classe (modifiable) :", value=recap_txt,
                                    height=280, key="recap_classe")
        r1, r2 = st.columns([2, 1])
        with r1:
            st.download_button("⬇️ Télécharger (.txt)", data=edited_recap.encode("utf-8"),
                file_name="recap_convocations_classe.txt", mime="text/plain", key="dl_recap")
        with r2:
            st.markdown(copy_button_html(edited_recap, "recap_classe"), unsafe_allow_html=True)

    st.markdown("---")

    # ── Mails individuels ──────────────────────────────────────────────────────
    st.markdown('<div class="section-title">📧 Mails individuels</div>', unsafe_allow_html=True)

    if filtered_df.empty:
        st.info("Aucun étudiant sélectionné — ajustez les filtres dans le panneau latéral.")
    else:
        to_convoke = []
        for _, row in filtered_df.iterrows():
            skey = f"{row['Prénom']} {row['Nom']}"
            mats = [
                col for col in eval_disp_cols
                if str(row.get(col,"")).strip() in RAT_VALS
                and not (use_comp and is_compensated_for(skey, col, student_ue_results))
            ]
            if mats:
                to_convoke.append((row["Prénom"], row["Nom"], mats))

        if not to_convoke:
            st.success("✅ Aucun étudiant avec C ou D non compensé.")
        else:
            st.markdown(f'<span class="stat-pill" style="margin-bottom:0.8rem;display:inline-flex;">📧 {len(to_convoke)} mail(s) à envoyer</span>', unsafe_allow_html=True)
            for prenom, nom, mats in to_convoke:
                with st.expander(f"📧 {prenom} {nom} — {len(mats)} matière(s)"):
                    badges = " ".join(
                        f'<span style="background:#fee2e2;color:#7f1d1d;border:1px solid #fca5a5;'
                        f'border-radius:6px;padding:2px 8px;font-size:0.78rem;font-weight:600;">{m}</span>'
                        for m in mats)
                    st.markdown(f"**Matières :** {badges}", unsafe_allow_html=True)
                    mail_txt = generate_email(prenom, nom, mats, tutoyer)
                    edited   = st.text_area("Mail (modifiable) :", value=mail_txt, height=260,
                                            key=f"mail_{tkey}_{prenom}_{nom}")
                    d1, d2 = st.columns([2, 1])
                    with d1:
                        st.download_button("⬇️ Télécharger (.txt)",
                            data=edited.encode("utf-8"),
                            file_name=f"mail_{prenom}_{nom}.txt".replace(" ","_"),
                            mime="text/plain", key=f"dl_{tkey}_{prenom}_{nom}")
                    with d2:
                        st.markdown(copy_button_html(edited, f"{tkey}_{prenom}_{nom}"),
                                    unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# ─── TAB 4 : CRÉNEAUX PARALLÈLES ────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

with tab4:
    st.markdown(
        "<p style='font-size:0.85rem;color:#6b7280;margin-bottom:1rem;'>"
        "Deux matières sont <strong>compatibles</strong> si elles n'ont "
        "<strong>aucun élève en commun</strong> parmi les convoqués non compensés.</p>",
        unsafe_allow_html=True)

    # recap_rows peut être vide si tab2 n'a pas été visité → on le recalcule ici si besoin
    if not recap_rows and not filtered_df.empty:
        for col in eval_disp_cols:
            ec, ed, ea, ec_comp, ed_comp = [], [], [], [], []
            for _, row in filtered_df.iterrows():
                skey  = f"{row['Prénom']} {row['Nom']}"
                grade = str(row.get(col, "")).strip()
                if grade == "C":
                    if use_comp and is_compensated_for(skey, col, student_ue_results):
                        ec_comp.append(skey)
                    else:
                        ec.append(skey)
                elif grade == "D":
                    if use_comp and is_compensated_for(skey, col, student_ue_results):
                        ed_comp.append(skey)
                    else:
                        ed.append(skey)
                elif grade == "ABS":
                    ea.append(skey)
            total_rat = len(ec) + len(ed) + len(ea)
            total_all = total_rat + len(ec_comp) + len(ed_comp)
            if total_all > 0:
                recap_rows.append({
                    "Matière": col,
                    "eleves_c": ec, "nb_c": len(ec),
                    "eleves_d": ed, "nb_d": len(ed),
                    "eleves_abs": ea, "nb_abs": len(ea),
                    "eleves_c_comp": ec_comp, "eleves_d_comp": ed_comp,
                    "total_rat": total_rat, "total_all": total_all,
                })

    if not recap_rows:
        st.info("Aucune donnée disponible — importez un fichier et appliquez les filtres.")
    else:
        mat_stu = {
            r["Matière"]: set(r["eleves_c"] + r["eleves_d"] + r["eleves_abs"])
            for r in recap_rows if r["total_rat"] > 0
        }
        mlist = list(mat_stu.keys())

        if len(mlist) < 2:
            st.info("Pas assez de matières pour calculer des compatibilités.")
        else:
            def compat(m1, m2): return mat_stu[m1].isdisjoint(mat_stu[m2])

            groupes, reste = [], list(mlist)
            while reste:
                g = [reste[0]]
                for m in reste[1:]:
                    if all(compat(m, gm) for gm in g):
                        g.append(m)
                groupes.append(g)
                reste = [m for m in reste if m not in g]

            COLORS = [
                ("#ede9fe","#4f46e5","#c4b5fd"), ("#d1fae5","#065f46","#6ee7b7"),
                ("#dbeafe","#1e3a8a","#93c5fd"), ("#fef3c7","#78350f","#fcd34d"),
                ("#fee2e2","#7f1d1d","#fca5a5"), ("#f3e8ff","#581c87","#d8b4fe"),
                ("#ffedd5","#7c2d12","#fb923c"),
            ]

            st.markdown(
                f"<p style='font-size:0.9rem;'><strong>{len(groupes)} créneau(x) minimum</strong> "
                f"nécessaire(s).</p>", unsafe_allow_html=True)

            for i, groupe in enumerate(groupes):
                bg2, fg2, bd2 = COLORS[i % len(COLORS)]
                all_elv = set()
                for m in groupe: all_elv |= mat_stu[m]
                lignes = ""
                for m in groupe:
                    elv_m  = sorted(mat_stu[m])
                    pills2 = "".join(
                        f'<span style="display:inline-block;background:white;color:{fg2};'
                        f'border:1px solid {bd2};border-radius:20px;padding:1px 9px;'
                        f'font-size:0.74rem;font-weight:600;margin:2px;">{e}</span>'
                        for e in elv_m)
                    lignes += (
                        f'<div style="margin-bottom:8px;">'
                        f'<span style="font-weight:700;font-size:0.85rem;">{m}</span>'
                        f'<span style="font-size:0.78rem;color:{fg2};opacity:0.8;margin-left:6px;">'
                        f'({len(elv_m)} élève(s))</span>'
                        f'<div style="margin-top:4px;">{pills2}</div></div>')
                st.markdown(
                    f'<div style="background:{bg2};border:1.5px solid {bd2};border-radius:12px;'
                    f'padding:14px 18px;margin-bottom:10px;box-shadow:0 2px 8px rgba(0,0,0,0.06);">'
                    f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">'
                    f'<span style="background:{fg2};color:white;border-radius:8px;padding:3px 12px;'
                    f'font-weight:800;font-size:0.9rem;">Créneau {i+1}</span>'
                    f'<span style="font-size:0.82rem;color:{fg2};font-weight:600;">'
                    f'{len(groupe)} matière(s) · {len(all_elv)} élève(s)</span></div>'
                    f'{lignes}</div>',
                    unsafe_allow_html=True)

            # ── Matrice de compatibilité avec noms en diagonale ───────────────
            with st.expander("🔍 Voir la matrice de compatibilité complète"):
                # En-têtes diagonaux avec les vrais noms (pas M1, M2...)
                th_cells = ""
                for j, m in enumerate(mlist):
                    # Nom court pour l'en-tête (max 30 chars)
                    label = m[:28] + "…" if len(m) > 30 else m
                    th_cells += (
                        f'<th style="background:#1e1b4b;color:white;padding:4px 2px;'
                        f'font-size:0.72rem;text-align:left;vertical-align:bottom;'
                        f'min-width:32px;max-width:32px;overflow:hidden;height:100px;">'
                        f'<div style="writing-mode:vertical-rl;transform:rotate(180deg);'
                        f'white-space:nowrap;font-weight:700;line-height:1.2;max-height:95px;'
                        f'overflow:hidden;text-overflow:ellipsis;padding:2px 0;">'
                        f'{label}</div></th>'
                    )
                header = (
                    f'<tr><th style="background:#1e1b4b;color:white;padding:7px 12px;'
                    f'font-size:0.78rem;text-align:left;white-space:nowrap;min-width:180px;">'
                    f'Matière</th>'
                    f'{th_cells}</tr>')

                body = ""
                for i_m, m1 in enumerate(mlist):
                    row_lbl = (
                        f'<td style="font-size:0.78rem;padding:5px 12px;white-space:nowrap;'
                        f'background:#f1f5f9;border-right:2px solid #e2e8f0;font-weight:600;">'
                        f'{m1}</td>')
                    cells2 = row_lbl
                    for j_m, m2 in enumerate(mlist):
                        if i_m == j_m:
                            cells2 += '<td style="background:#e2e8f0;text-align:center;font-size:0.75rem;">—</td>'
                        elif compat(m1, m2):
                            cells2 += '<td style="background:#d1fae5;color:#065f46;text-align:center;font-weight:800;font-size:0.85rem;">✓</td>'
                        else:
                            nb = len(mat_stu[m1] & mat_stu[m2])
                            cells2 += (f'<td style="background:#fee2e2;color:#7f1d1d;'
                                       f'text-align:center;font-weight:700;font-size:0.8rem;">{nb}</td>')
                    bg_r = "#f8fafc" if i_m % 2 == 0 else "white"
                    body += f'<tr style="background:{bg_r}">{cells2}</tr>'

                st.markdown(
                    '<p style="font-size:0.8rem;color:#6b7280;margin-bottom:6px;">'
                    '✓ = compatible · chiffre = nb d\'élèves en conflit</p>'
                    '<div style="overflow-x:auto;border-radius:10px;box-shadow:0 1px 6px rgba(0,0,0,0.08);">'
                    f'<table style="border-collapse:collapse;font-family:sans-serif;">'
                    f'<thead>{header}</thead><tbody>{body}</tbody></table></div>',
                    unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# ─── TAB 5 : EXPORT EXCEL ───────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

with tab5:
    st.markdown('<div class="section-title">📥 Export Excel</div>', unsafe_allow_html=True)

    onglets_desc = [
        ("📊 Vue globale", "Toutes les personnes × toutes les matières, avec code couleur mentions"),
        ("📋 Rattrapages", "Tous les étudiants : matières en rattrapage + matières compensées"),
        ("🔍 Mentions (filtrés)", "Étudiants correspondant aux filtres actifs"),
    ]
    if use_comp and student_ue_results:
        onglets_desc.append(("⚖️ Résultats UE", "Résultats par UE avec statut de compensation"))

    st.markdown("**Contenu de l'export :**")
    for icon_title, desc in onglets_desc:
        st.markdown(
            f'<div style="display:flex;align-items:flex-start;gap:10px;'
            f'padding:8px 12px;background:white;border-radius:8px;margin-bottom:5px;'
            f'border:1px solid #e8eaf6;">'
            f'<span style="font-weight:700;color:#1e1b4b;min-width:200px;">{icon_title}</span>'
            f'<span style="font-size:0.82rem;color:#6b7280;">{desc}</span></div>',
            unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    if not display_df.empty:
        lbl = "⬇️ Télécharger l'export Excel (.xlsx)"
        st.download_button(
            label=lbl,
            data=make_excel(filtered_df, eval_disp_cols, student_ue_results, use_comp, display_df),
            file_name="rattrapages_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
        st.caption("ℹ️ La Vue globale et l'onglet Rattrapages incluent TOUS les étudiants (pas seulement les filtrés).")
    else:
        st.warning("Aucune donnée à exporter.")
